import openpyxl
import calendar
import datetime  

def createCronogram(year: int,names: list, suc_names: list):
    file_path = "./excel/result.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    month = 1
    day_of_month = 1
    first_day_of_year = datetime.date(year, 1, 1)
    start_day = first_day_of_year.weekday()
    print(start_day)
    # Rotate the days_of_week array so it starts on the correct day
    days_of_week = ['Lu', 'Ma', 'Mi', 'Ju', 'Vi', 'Sa', 'Do']
    days_of_week = days_of_week[start_day:] + days_of_week[:start_day]

    
    months = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio',
              'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    num_names = len(names)
    
    current_week = -1
    
    
    
    rotations = [
        ['N', 'N', 'N', 'N', 'N', 'TC', 'TC'],  # Turn week
        ['D', 'D', 'D', 'D', 'D', 'L', 'L'],  # Not turn week
        ['A', 'A', 'A', 'A', 'A', 'TC', 'TC'],  # Turn week
        ['D', 'D', 'D', 'D', 'D', 'L', 'L'],  # Not turn week
        ['D', 'D', 'D', 'D', 'D', 'L', 'L']   # Not turn week
    ]
    
    suc_rotations = [
        ['D', 'D', 'D', 'D', 'D', 'TC', 'TC'],  # Turn week
        ['D', 'D', 'D', 'D', 'D', 'L', 'L'],  # Not turn week
        ['D', 'D', 'D', 'D', 'D', 'TC', 'TC'],  # Turn week
        ['D', 'D', 'D', 'D', 'D', 'L', 'L'],  # Not turn week
        ['D', 'D', 'D', 'D', 'D', 'L', 'L']   # Not turn week
    ]
    for day in range(1, 367):  # year is a leap year, so 366 days
        if day_of_month > calendar.monthrange(year, month)[1]:
            month += 1
            day_of_month = 1
            if month > 12:
                break

        date = datetime.date(year, month, day_of_month)
        day_of_week = days_of_week[date.weekday()]
        week_of_year = date.isocalendar()[1]
        
        if date.weekday() == 1:
            week_of_year += 1
        
        
        cell_day = sheet.cell(row=3, column=day + 1)
        cell_day.value = day_of_month

        cell_weekday = sheet.cell(row=2, column=day + 1)
        cell_weekday.value = day_of_week

        cell_label = sheet.cell(row=2, column=1)
        cell_label.value = "Día"
        
        cell_label = sheet.cell(row=3, column=1)
        cell_label.value = "Nombre"
        
        if day_of_week == 'Ma':
            current_week = (current_week + 1) % num_names

        rotate(names, rotations, current_week, date, day, sheet)
        rotate(suc_names, suc_rotations, current_week, date, day, sheet, True)
            
        day_of_month += 1
        
      
    
    sheet.insert_rows(1)
    monthspassed = 0
    for column in sheet.iter_cols(min_col=2, max_col=sheet.max_column):
        if column[3].value == 1:
            column[1].value = months[monthspassed]
            monthspassed += 1
    sheet.delete_rows(1)            
    workbook.save(file_path)

def rotate(names: list, rotations: list, current_week: int, date: datetime.date, day: int, sheet: openpyxl.worksheet.worksheet.Worksheet, isExtending: bool = False):
    for i, name in enumerate(names, start=4 if not isExtending else 10):
            rotation_index = (current_week + i) % len(rotations)
            rotation = rotations[rotation_index]
            
            cell = sheet.cell(row=i, column=day + 1)
            cell.value = rotation[date.weekday()]
            
    for i, name in enumerate(names, start=4 if not isExtending else 10):
        cell_name = sheet.cell(row=i, column=1)
        cell_name.value = name        