from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment

tborder = Side(border_style="thin")
sborder = Side(border_style="thick")


def ConstructTables(table):
    ws = table
    Table = []
    for row in ws.iter_rows():
        Row = []
        for cell in row:
            Row.append(cell)
        Table.append(Row)
    return Table


def Color():
    wb = load_workbook("test.xlsx")
    ws = wb.active
    table = ConstructTables(ws)
    for row in table:
        for cell in row:
            value = cell.value
            if value:
                ws[cell.coordinate].border = Border(
                    top=tborder, bottom=tborder, left=tborder, right=tborder)
                ws[cell.coordinate].alignment = Alignment(shrink_to_fit=True)
            if value in ["Sab", "Dom", "L"]:
                ws[cell.coordinate].fill = PatternFill(
                    fill_type="solid", start_color="33CCCC")
                DayCoord = cell.coordinate[:-1] + "4"
                ws[DayCoord].fill = PatternFill(
                    fill_type="solid", start_color="33CCCC")
            elif value == "D":
                ws[cell.coordinate].fill = PatternFill(
                    fill_type="solid", start_color="FFFF00")
            elif value in ["N", "TC", "TP"]:
                ws[cell.coordinate].fill = PatternFill(
                    fill_type="solid", start_color="000000")
                ws[cell.coordinate].font = Font(color="FFFFFF")
            elif value == "A":
                ws[cell.coordinate].fill = PatternFill(
                    fill_type="solid", start_color="8497B0")
                ws[cell.coordinate].font = Font(color="FFFFFF")
            elif cell.coordinate[-1] == "9":
                ws[cell.coordinate].fill = PatternFill(
                    fill_type="solid", start_color="E7E6E6")

    wb.save("test.xlsx")
