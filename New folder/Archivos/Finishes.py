from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment


def ConstructTables(table):
    ws = table
    Table = []
    for row in ws.iter_rows():
        Row = []
        for cell in row:
            Row.append(cell)
        Table.append(Row)
    return Table

def Finish(TurnoList, CodigoList, TrabajadorList):
    wb = load_workbook("test.xlsx")
    ws = wb.active
    ws.insert_cols(0)
    table = ConstructTables(ws)
    ws["A4"],ws["B4"],ws["C4"] = "TURNO","CODIGO","NOMBRE"
    for row in table:
        for cell in row:
            try:
                if row[row.index(cell) + 1].value == "Código" :
                    cell.value = "Turno"
            except:
                pass
    wb.save("test.xlsx")
    table = ConstructTables(ws)
    i,j,k = 0,0,0
    for row in table:
        for cell in row:
            match cell.value:

                case "Turno":
                    cell.value = TurnoList[i]
                    i += 1
                case "Código":
                    cell.value = CodigoList[j]
                    j += 1
                case "Nombre":
                    cell.value = TrabajadorList[k]
                    k += 1
    wb.save("test.xlsx")
